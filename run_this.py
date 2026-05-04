import tensorflow as tf
import numpy as np
from openpyxl import Workbook
import time
import ddpg_brain
from env import UAVEnv
from state_normalization import StateNormalization
############################################################################################################################

# 创建每回合奖励文件
file_path1 = 'ddpg奖励函数episodes3.xlsx'
workbook1 = Workbook()
sheet1 = workbook1.active
sheet1.append(["每回合奖励函数"])

#创建每回合成功任务数
file_path5 = 'ddpg每回合成功任务数episodes3.xlsx'
workbook5 = Workbook()
sheet5 = workbook5.active
sheet5.append(["每回合成功任务数"])

# 创建每回合所有终端能耗文件
file_path4 = 'ddpg每个回合所有终端能耗ep3.xlsx'
workbook4 = Workbook()
sheet4 = workbook4.active
sheet4.append(["每回合终端能耗"])

#创建每回合无人机能耗
file_path3 = 'ddpg每个回合无人机的能耗episodes3.xlsx'
workbook3 = Workbook()
sheet3 = workbook3.active
sheet3.append(["每回合无人机能耗"])

#创建每回合成功卸载到无人机的任务数
file_path6 = 'ddpg每个回合成功卸载到无人机的任务数episodes3.xlsx'
workbook6 = Workbook()
sheet6 = workbook6.active
sheet6.append(["每回合成功卸载到无人机的任务数"])

#创建每回合成功卸载到卫星的任务数
file_path7 = 'ddpg每个回合成功卸载到卫星的任务数episodes3.xlsx'
workbook7 = Workbook()
sheet7 = workbook7.active
sheet7.append(["每回合成功卸载到卫星的任务数"])

#######################################################################################################################
#用具体的环境训练ddpg智能体的参数设置
np.random.seed(1)
tf.set_random_seed(1)
np.set_printoptions(suppress=True)  # 禁止科学计数法
MAX_EPISODES=550
MAX_EP_STEPS = 150#每个回合的最大步数150
env = UAVEnv()#具体的环境
a_bound = env.action_bound#[-1,1]裁剪动作空间
s_dim = env.state_dim#状态空间的维数
a_dim = env.action_dim#动作空间的维数
ddpg = ddpg_brain.DDPG(a_dim, s_dim, a_bound)#ddpg智能体类的一个实例
var = 0.5#探索因子
t1 = time.time()#程序开始执行的时刻
s_normal = StateNormalization()#对状态空间标准化的一个实例
################################################################################################################
#用具体的回合环境训练ddpg智能体
for episode in range(MAX_EPISODES):
    print(f"一个大的回合开始：{episode}！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！！")
    ep_reward_func= 0#每个回合的奖励函数列表
    ep_success_num=0#每个回合成功任务数
    ep_ue_e_consume=0#每个回合所有终端的能耗
    ep_UAV_energy_consume=0#每个回合无人机的能耗
    ep_offloading_uav=0#每个回合成功卸载到无人机的任务数
    ep_offloading_satellite=0#每个回合成功卸载到卫星的任务数
    # ep_delay=0
    s = env.reset()#这个表示每个回合，每个回合开始前重置所有环境
    print(f"一个回合开始的初始状态:"
          f"\n任务大小：{s[0:UAVEnv.M]}\n终端电量：{s[UAVEnv.M:2*UAVEnv.M]}"
          f"\n无人机电量：{s[2*UAVEnv.M]}\n每个终端的位置：{s[2*UAVEnv.M+1:4*UAVEnv.M+1]}"
          f"\n无人机的位置：{s[4*UAVEnv.M+1:4*UAVEnv.M+3]}\n遮挡标记：{s[4*UAVEnv.M+3:5*UAVEnv.M+3]}"
          f"\n任务截止时间：{s[5*UAVEnv.M+3:6*UAVEnv.M+3]}")
    for j in range(MAX_EP_STEPS):
        print(f"第{episode}回合第{j}个时间片开始@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@")
        a = ddpg.choose_action(s_normal.state_normal(s))
        var = max([var * 0.9997, ddpg_brain.VAR_MIN])  # decay the action randomness
        a = np.clip(np.random.normal(a, var), *a_bound)#为动作添加均值为a，标准差为var的高斯噪声
        print(a)
        s_next,reward_func,success_num,ue_e_consume,uav_e_consume,offloading_uav,offloading_satellite,done=env.step(a)#执行动作a后环境的状态
        ddpg.store_transition(s_normal.state_normal(s), a, reward_func, s_normal.state_normal(s_next))
        #如果存储的经验比经验回放池大就用经验进行训练
        # if ddpg.pointer > ddpg_brain_sigmoid.MEMORY_CAPACITY and ddpg.pointer%100:
        if ddpg.pointer > ddpg_brain.MEMORY_CAPACITY:
            ddpg.learn()
        s = s_next
        print("下一个状态"f"\n任务大小：{s[0:UAVEnv.M]}\n终端电量：{s[UAVEnv.M:2*UAVEnv.M]}"
          f"\n无人机电量：{s[2*UAVEnv.M]}\n每个终端的位置：{s[2*UAVEnv.M+1:4*UAVEnv.M+1]}"
          f"\n无人机的位置：{s[4*UAVEnv.M+1:4*UAVEnv.M+3]}\n遮挡标记：{s[4*UAVEnv.M+3:5*UAVEnv.M+3]}"
          f"\n任务截止时间：{s[5*UAVEnv.M+3:6*UAVEnv.M+3]}")
        #累计性能信息####################################################################################################
        ep_reward_func+=reward_func#累计每个回合的奖励函数
        ep_success_num+=success_num#累计每回合的奖励
        ep_ue_e_consume+=ue_e_consume#累计每回合所有终端的能耗
        ep_UAV_energy_consume+=uav_e_consume#累计每回合无人机的能耗
        ep_offloading_uav+=offloading_uav#累计每回合无人机的能耗
        ep_offloading_satellite+=offloading_satellite#累计每回合无人机的能耗
        #####################################################################################################累计性能信息#
        if j == MAX_EP_STEPS - 1 or done:#如果达到最大时间步或者结束条件
            #向表格中添加信息#########################################################################################################
            sheet1.append([ep_reward_func])#向表格中添加每回合奖励函数
            sheet5.append([ep_success_num])#向表格中添加每回合奖励
            sheet4.append([ep_ue_e_consume])#向表格中添加终端能耗
            sheet3.append([ep_UAV_energy_consume])#向表格中添加无人机能耗
            sheet6.append([ep_offloading_uav])#向表格中添加成功卸载到无人机的任务数
            sheet7.append([ep_offloading_satellite])#向表格中添加成功卸载到卫星的任务数
            break
    print(f"一个大的回合第{episode}结束")
#########################################################################################################向表格中添加信息#
# ddpg.save_model('ddpg_model/model_episodes_{}.ckpt'.format(name))
#保存表格###############################################################################################################
# 保存Excel文件
workbook1.save(file_path1)#存储奖励函数文件
workbook5.save(file_path5)#存储每回合成功任务数
workbook4.save(file_path4)#存储所有终端能耗文件
workbook3.save(file_path3)#存储无人机能耗文件
workbook6.save(file_path6)#存储成功卸载到无人机的任务数
workbook7.save(file_path7)#存储成功卸载到卫星的任务数
# workbook2.save(file_path2)#存储延迟文件
###############################################################################################################保存表格#
#记录程序运行时间
print('Running time: ', time.time() - t1)



# print("接下来是验证")
# # 加载训练好的模型
# tf.reset_default_graph()
# ddpg = DDPG(a_dim, s_dim, a_bound)  # 重新初始化 DDPG
# ddpg.load_model('ddpg_model/model_episode_m5.ckpt')
#
# print("模型参数已加载成功！")
# # 加载模型后可以继续进行训练或测试
# #记录性能信息############################################################################################################
# # 创建每回合奖励文件
# file_path5 = r'训练好的ddpg成功任务数.xlsx'
# workbook5 = Workbook()
# sheet5 = workbook5.active
# sheet5.append(["每回合成功任务数"])
#
# # # 创建每回合延迟文件
# # file_path2 = r'训练好的ddpg延迟.xlsx'
# # workbook2 = Workbook()
# # sheet2 = workbook2.active
# # sheet2.append(["每回合任务延迟"])
#
# # # 创建每回合无人机能耗文件
# # file_path3 = r'训练好的ddpg无人机能耗.xlsx'
# # workbook3 = Workbook()
# # sheet3 = workbook3.active
# # sheet3.append(["每回合无人机能耗"])
#
# # 创建每回合无人机能耗文件
# file_path6 = r'训练好的ddpg终端能耗.xlsx'
# workbook6 = Workbook()
# sheet6 = workbook6.active
# sheet6.append(["每回合无人机能耗"])
# ############################################################################################################记录性能信息#
#
# t2 = time.time()
# for i in range(10):
#     s = env.reset()
#     ep_success_num=0
#     # ep_delay=0
#     # ep_UAV_energy_consume=0
#     ep_ue_energy_consume=0
#
#     j = 0
#     for t in range(MAX_EP_STEPS):
#         a = ddpg.choose_action(s_normal.state_normal(s))
#         a = np.clip(np.random.normal(a, var), *a_bound)#为动作添加均值为a，标准差为var的高斯噪声
#         s_,R_func,success,ue_e_consume,done=env.step(a)#执行动作a后环境的状态
#         s = s_
#         #累计性能信息####################################################################################################
#         ep_success_num+=success#累计每回合的奖励
#         # ep_delay+=task_total_delay_step#累计每回合延迟
#         # ep_UAV_energy_consume+=UAV_energy_consume_step#累计每回合无人机能耗
#         ep_ue_energy_consume+=ue_e_consume#累计每回合无人机能耗
#         #####################################################################################################累计性能信息#
#
#         # if reset_dist:  # 如果无人机越界
#         #
#         #     break
#         #
#         # if eu_lack_energy:  # 如果终端没能量了
#         #     break
#         #
#         # if UAV_lack_energy:  # 如果无人机没能量了
#         #     break
#         if t == MAX_EP_STEPS - 1 or done:
#             # 向表格中添加信息#########################################################################################################
#             sheet5.append([ep_success_num])  # 向表格中添加每回合成功任务数
#             # sheet2.append([ep_delay])#向表格中添加延迟
#             # sheet3.append([ep_UAV_energy_consume])#向表格中添加无人机能耗
#             sheet6.append([ep_ue_energy_consume])  # 向表格中添加终端能耗
#             # file_name = 'output_ddpg_' + str(self.bandwidth_nums) + 'MHz.txt'
#             break
#
#     #向表格中添加信息#########################################################################################################
#     # sheet1.append([ep_success_num])#向表格中添加奖励
#     # sheet2.append([ep_delay])#向表格中添加延迟
#     # sheet3.append([ep_UAV_energy_consume])#向表格中添加无人机能耗
# #########################################################################################################向表格中添加信息#
# #记录程序运行时间
# print('Running time: ', time.time() - t2)
# #保存表格###############################################################################################################
# # 保存Excel文件
# workbook5.save(file_path5)#存储奖励文件
# print("已保存5")
# # workbook2.save(file_path2)#存储延迟文件
# # workbook3.save(file_path3)#存储无人机能耗文件
# workbook6.save(file_path6)#存储无人机能耗文件
# print("已保存6")

###############################################################################################################保存表格#
